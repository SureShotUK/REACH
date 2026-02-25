#!/usr/bin/env python3
"""
Build maintenance admin workbooks:
  - Compliance_Schedule.xlsx
  - Job_Log.xlsx

Run with:  python3 build_workbooks.py
Requires:  openpyxl >= 3.1
"""

import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ── Colour palette ─────────────────────────────────────────────────────────────

def _fill(hex_col: str) -> PatternFill:
    return PatternFill(start_color=hex_col, end_color=hex_col, fill_type="solid")

NAVY         = _fill("1F3864")   # header row
BLUE_LIGHT   = _fill("D9E1F2")   # alternate data row
WHITE_FILL   = _fill("FFFFFF")

# Status fills
F_OVERDUE    = _fill("FF0000")   # red
F_DUE30      = _fill("FF4B00")   # dark orange
F_DUE90      = _fill("FFC000")   # amber
F_CURRENT    = _fill("92D050")   # green
F_NOTSCHED   = _fill("BFBFBF")   # grey

# Job-status fills
F_OPEN       = _fill("BDD7EE")   # pale blue
F_INPROG     = _fill("FFE699")   # pale amber
F_PEND_INV   = _fill("F4B942")   # orange-amber
F_CLOSED     = _fill("C6EFCE")   # pale green
F_CANCELLED  = _fill("E0E0E0")   # light grey

HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT    = Font(name="Calibri", size=10)
STATUS_BOLD  = Font(name="Calibri", bold=True, size=10)

WRAP_TOP  = Alignment(wrap_text=True, vertical="top")
CENTRE    = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_TOP  = Alignment(horizontal="left", vertical="top", wrap_text=True)

def _thin_border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


# ── Generic helpers ────────────────────────────────────────────────────────────

def _set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _style_header(ws, row: int, n_cols: int):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = NAVY
        cell.font = HEADER_FONT
        cell.alignment = CENTRE
        cell.border = _thin_border()


def _style_data_rows(ws, start_row: int, end_row: int, n_cols: int):
    for r in range(start_row, end_row + 1):
        fill = BLUE_LIGHT if r % 2 == 0 else WHITE_FILL
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.font = BODY_FONT
            cell.alignment = WRAP_TOP
            cell.border = _thin_border()


def _add_dv_list(ws, formula1: str, sqref: str, prompt: str = "", title: str = ""):
    dv = DataValidation(
        type="list",
        formula1=formula1,
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid entry",
        error="Please select a value from the dropdown list.",
        showInputMessage=bool(prompt),
        promptTitle=title,
        prompt=prompt,
    )
    ws.add_data_validation(dv)
    dv.sqref = sqref


def _add_instructions_sheet(wb: Workbook, title: str, lines: list):
    ws = wb.create_sheet(title)
    ws.column_dimensions["A"].width = 110
    ws.cell(1, 1).value = title
    ws.cell(1, 1).font = Font(name="Calibri", bold=True, size=14, color="1F3864")
    ws.cell(1, 1).alignment = LEFT_TOP
    ws.row_dimensions[1].height = 24
    for i, line in enumerate(lines, start=2):
        cell = ws.cell(i, 1)
        cell.value = line
        cell.font = BODY_FONT
        cell.alignment = LEFT_TOP
        ws.row_dimensions[i].height = 15 if line else 8


# ══════════════════════════════════════════════════════════════════════════════
#  COMPLIANCE SCHEDULE
# ══════════════════════════════════════════════════════════════════════════════

COMP_HEADERS = [
    "Record ID",              # A
    "Site",                   # B
    "Asset ID",               # C
    "Asset / System",         # D
    "Location within Site",   # E
    "Regulation / Standard",  # F
    "Requirement",            # G
    "Category",               # H
    "Frequency",              # I
    "Interval (Days)",        # J  0 = risk-assessed/manual
    "Last Completed",         # K  date entered by user
    "Certificate / Report Ref",  # L
    "Next Due Date",          # M  formula (or manual for risk-assessed)
    "Days Until Due",         # N  formula
    "Status",                 # O  formula
    "Responsible Contractor", # P
    "Notes / Actions",        # Q
]

COMP_COL_WIDTHS = {
    "A": 12, "B":  8, "C": 16, "D": 30, "E": 22,
    "F": 32, "G": 38, "H": 10, "I": 22, "J": 12,
    "K": 15, "L": 24, "M": 15, "N": 14, "O": 22,
    "P": 26, "Q": 44,
}

# Tuple: (record_id, site, asset_id, asset_system, location,
#         regulation, requirement, category, frequency, interval_days, notes)
COMP_DATA = [
    # ── CITY ─────────────────────────────────────────────────────────────────
    (
        "COMP-0001", "CITY", "CITY-GAS-001", "Gas Boiler", "Boiler Room / Plant Room",
        "Gas Safety (Installation & Use) Regulations 1998",
        "Annual gas boiler service by Gas Safe registered engineer",
        "STAT", "Annual", 365, "",
    ),
    (
        "COMP-0002", "CITY", "CITY-ELEC-001", "Fixed Electrical Installation", "All Areas",
        "Electricity at Work Regulations 1989",
        "Electrical Installation Condition Report (EICR)",
        "STAT", "Every 5 years (offices)", 1825, "",
    ),
    (
        "COMP-0003", "CITY", "CITY-FIRE-001", "Fire Alarm System", "All Areas",
        "Regulatory Reform (Fire Safety) Order 2005",
        "Weekly fire alarm test by nominated Responsible Person",
        "STAT", "Weekly", 7,
        "Test must be recorded in the fire log book. Rotate zones tested.",
    ),
    (
        "COMP-0004", "CITY", "CITY-FIRE-001", "Fire Alarm System", "All Areas",
        "Regulatory Reform (Fire Safety) Order 2005",
        "Six-monthly fire alarm inspection and service by competent fire engineer",
        "STAT", "Every 6 months", 182, "",
    ),
    (
        "COMP-0005", "CITY", "CITY-FIRE-002", "Emergency Lighting", "All Areas",
        "Regulatory Reform (Fire Safety) Order 2005 / BS 5266-1",
        "Monthly emergency lighting function test (flick test)",
        "STAT", "Monthly", 30,
        "Record results in lighting log book.",
    ),
    (
        "COMP-0006", "CITY", "CITY-FIRE-002", "Emergency Lighting", "All Areas",
        "Regulatory Reform (Fire Safety) Order 2005 / BS 5266-1",
        "Annual full rated-duration emergency lighting test (3-hour test)",
        "STAT", "Annual", 365,
        "Plan during low-occupancy period. All luminaires must sustain full rated duration.",
    ),
    (
        "COMP-0007", "CITY", "CITY-FIRE-003", "Fire Extinguishers", "All Areas",
        "Regulatory Reform (Fire Safety) Order 2005 / BS 5306-3",
        "Annual service and inspection of all fire extinguishers by competent person",
        "STAT", "Annual", 365, "",
    ),
    (
        "COMP-0008", "CITY", "CITY-HVAC-001", "Air Conditioning Unit(s)", "Office Areas",
        "F-Gas Regulation (EU) 517/2014 (retained in UK law)",
        "F-Gas refrigerant leak check - frequency determined by CO2e charge size",
        "STAT", "Risk-assessed - confirm CO2e charge with F-Gas engineer", 0,
        "ACTION REQUIRED: confirm refrigerant type and CO2e charge for each unit. "
        "<5 tCO2e = annual check; 5-50 = 6-monthly; >50 = quarterly. "
        "Only F-Gas certified engineer may work on system.",
    ),
    (
        "COMP-0009", "CITY", "CITY-HVAC-001", "Air Conditioning Unit(s)", "Office Areas",
        "PPM - Manufacturer / Best Practice",
        "Annual service, filter clean and performance check",
        "PPM", "Annual", 365, "",
    ),
    (
        "COMP-0010", "CITY", "CITY-L8-001", "Water System (Cold & Hot)", "All Areas",
        "L8 ACoP / HSG274 Part 2 - Legionella Control",
        "Monthly hot and cold water temperature monitoring at sentinel taps",
        "STAT", "Monthly", 30,
        "Hot water: 50°C+ at outlets within 1 min; 60°C at calorifier. "
        "Cold water: below 20°C at outlets. Investigate and record any exceedances immediately.",
    ),
    (
        "COMP-0011", "CITY", "CITY-L8-001", "Water System (Cold & Hot)", "All Areas",
        "L8 ACoP / HSG274 - Legionella Control",
        "Legionella risk assessment review by competent person",
        "STAT", "Every 2 years", 730,
        "Review sooner if significant changes to the building or water system occur.",
    ),
    (
        "COMP-0012", "CITY", "CITY-ELEC-002", "Portable Appliances", "All Areas",
        "Electricity at Work Regulations 1989",
        "PAT testing - frequency set by formal risk assessment",
        "STAT", "Risk-assessed", 0,
        "ACTION REQUIRED: complete PAT risk assessment to set formal frequency. "
        "Offices typically 2-4 years for Class I equipment; annually for Class II.",
    ),

    # ── MFG ──────────────────────────────────────────────────────────────────
    (
        "COMP-0013", "MFG", "MFG-ELEC-001", "Fixed Electrical Installation", "All Areas",
        "Electricity at Work Regulations 1989",
        "Electrical Installation Condition Report (EICR) - Industrial premises",
        "STAT", "Every 3 years (industrial)", 1095, "",
    ),
    (
        "COMP-0014", "MFG", "MFG-FIRE-001", "Fire Alarm System", "All Areas",
        "Regulatory Reform (Fire Safety) Order 2005",
        "Weekly fire alarm test by nominated Responsible Person",
        "STAT", "Weekly", 7,
        "Record in fire log book. Rotate zones tested each week.",
    ),
    (
        "COMP-0015", "MFG", "MFG-FIRE-001", "Fire Alarm System", "All Areas",
        "Regulatory Reform (Fire Safety) Order 2005",
        "Six-monthly fire alarm inspection and service by competent fire engineer",
        "STAT", "Every 6 months", 182, "",
    ),
    (
        "COMP-0016", "MFG", "MFG-FIRE-002", "Emergency Lighting", "All Areas incl. Warehouse",
        "Regulatory Reform (Fire Safety) Order 2005 / BS 5266-1",
        "Monthly emergency lighting function test",
        "STAT", "Monthly", 30,
        "Record results in lighting log book. Include warehouse and portakabin areas.",
    ),
    (
        "COMP-0017", "MFG", "MFG-FIRE-002", "Emergency Lighting", "All Areas incl. Warehouse",
        "Regulatory Reform (Fire Safety) Order 2005 / BS 5266-1",
        "Annual full rated-duration emergency lighting test (3-hour test)",
        "STAT", "Annual", 365, "",
    ),
    (
        "COMP-0018", "MFG", "MFG-FIRE-003", "Fire Extinguishers", "All Areas",
        "Regulatory Reform (Fire Safety) Order 2005 / BS 5306-3",
        "Annual service of all fire extinguishers by competent person",
        "STAT", "Annual", 365, "",
    ),
    (
        "COMP-0019", "MFG", "MFG-HVAC-001", "Air Conditioning Unit(s)", "Offices / Workshop",
        "F-Gas Regulation (EU) 517/2014 (retained in UK law)",
        "F-Gas refrigerant leak check - frequency by CO2e charge",
        "STAT", "Risk-assessed - confirm CO2e charge with F-Gas engineer", 0,
        "ACTION REQUIRED: confirm refrigerant type and CO2e charge for each unit. "
        "<5 tCO2e = annual; 5-50 = 6-monthly; >50 = quarterly.",
    ),
    (
        "COMP-0020", "MFG", "MFG-HVAC-001", "Air Conditioning Unit(s)", "Offices / Workshop",
        "PPM - Manufacturer / Best Practice",
        "Annual service, filter clean and performance check",
        "PPM", "Annual", 365, "",
    ),
    (
        "COMP-0021", "MFG", "MFG-L8-001",
        "Water System incl. Shower Facilities", "Portakabin / Offices",
        "L8 ACoP / HSG274 Part 2 - Legionella Control",
        "Monthly hot and cold water temperature monitoring at sentinel taps incl. shower feeds",
        "STAT", "Monthly", 30,
        "** HIGH RISK - SHOWERS PRESENT ** Portakabin water systems may have low thermal "
        "mass and poor insulation. Hot water must reach 60°C at calorifier and 50°C at outlets "
        "within 1 minute. Cold water must be below 20°C. Record all results; escalate exceedances immediately.",
    ),
    (
        "COMP-0022", "MFG", "MFG-L8-002", "Showerheads and Hoses", "Shower Facilities",
        "L8 ACoP / HSG274 Part 2 - Legionella Control",
        "Quarterly removal, descaling, disinfection and cleaning of all showerheads and hoses",
        "STAT", "Quarterly", 91,
        "** HIGH RISK ** Showerheads are a primary Legionella aerosol risk. "
        "Remove head, soak in disinfectant (e.g. 50ppm chlorine solution for 1 hour), "
        "rinse and refit. Replace if descaling is inadequate. Record disinfection method and result.",
    ),
    (
        "COMP-0023", "MFG", "MFG-L8-001",
        "Water System incl. Shower Facilities", "Portakabin / Offices",
        "L8 ACoP / HSG274 - Legionella Control",
        "Legionella risk assessment by competent person (showers present - higher risk)",
        "STAT", "Every 2 years", 730,
        "** HIGH RISK ** Shower facilities elevate this site above standard office risk. "
        "Risk assessment must specifically cover shower system design, water temperatures, "
        "and portakabin water system characteristics. Review sooner if system changes.",
    ),
    (
        "COMP-0024", "MFG", "MFG-L8-001",
        "Infrequently Used Water Outlets", "All Areas",
        "L8 ACoP / HSG274 - Legionella Control",
        "Regular flushing of any outlet not used in the previous 7 days",
        "PPM", "Weekly (for unused outlets)", 7,
        "Any tap, shower or outlet not used for 7+ days must be flushed for at least 2 minutes "
        "before being returned to normal use. Record outlets flushed and date.",
    ),
    (
        "COMP-0025", "MFG", "MFG-LIFT-001",
        "Forklift Truck(s) - Main Body", "Warehouse",
        "Lifting Operations and Lifting Equipment Regulations (LOLER) 1998",
        "Thorough examination of FLT(s) by competent person",
        "STAT", "Every 12 months (standard FLT)", 365,
        "Examination report (form F97 or equivalent) must be retained. "
        "If FLT is used to lift persons (e.g. with man-basket), interval reduces to 6 months. "
        "Add a separate record per FLT if multiple trucks are on site.",
    ),
    (
        "COMP-0026", "MFG", "MFG-LIFT-002",
        "FLT Forks and Lifting Attachments", "Warehouse",
        "Lifting Operations and Lifting Equipment Regulations (LOLER) 1998",
        "Thorough examination of lifting accessories (forks, chains, slings, shackles, hooks)",
        "STAT", "Every 6 months", 182,
        "Includes all loose lifting gear used with the FLT. "
        "Examination records must be retained. Remove any item that fails examination from service.",
    ),
    (
        "COMP-0027", "MFG", "MFG-PRES-001",
        "Compressed Air System", "Workshop / Production",
        "Pressure Systems Safety Regulations (PSSR) 2000",
        "Thorough examination per Written Scheme of Examination by competent person",
        "STAT", "Per Written Scheme (typically 12-26 months for receiver vessels)", 0,
        "** ACTION REQUIRED ** Confirm Written Scheme of Examination (WSE) is in place. "
        "Interval is set by the competent person in the WSE. "
        "DO NOT operate pressure system without a current, valid Written Scheme.",
    ),
    (
        "COMP-0028", "MFG", "MFG-PRES-001",
        "Compressed Air System - Written Scheme", "Workshop / Production",
        "Pressure Systems Safety Regulations (PSSR) 2000",
        "Written Scheme of Examination in place, current and covering all relevant parts",
        "STAT", "Review when system changes", 0,
        "** ACTION REQUIRED ** Verify WSE exists, is held on site, and covers all "
        "pressure-containing components (receiver vessel, pipework, safety valves). "
        "Must be prepared or certified by a competent person.",
    ),
    (
        "COMP-0029", "MFG", "MFG-RACK-001",
        "Warehouse Racking", "Warehouse",
        "SEMA Code of Practice / Provision and Use of Work Equipment Regs (PUWER) 1998",
        "Annual expert inspection by SEMA-accredited racking inspector",
        "STAT", "Annual", 365,
        "Inspection must be by a SEMA-accredited inspector. "
        "Obtain written report with risk ratings (green/amber/red). "
        "Act on amber and red items within timescales specified in report. "
        "Also required after any collision or damage incident.",
    ),
    (
        "COMP-0030", "MFG", "MFG-RACK-001",
        "Warehouse Racking", "Warehouse",
        "SEMA Code of Practice / PUWER 1998",
        "Regular user racking inspection by trained and appointed Rack Safety Officer",
        "PPM", "Monthly (minimum)", 30,
        "Appoint a Rack Safety Officer and provide SEMA-recognised training. "
        "Use structured inspection checklist. Document all inspections. "
        "Remove damaged beams/uprights from service immediately - do not wait for expert inspection.",
    ),
    (
        "COMP-0031", "MFG", "MFG-ELEC-002",
        "Portable Appliances", "All Areas",
        "Electricity at Work Regulations 1989",
        "PAT testing - frequency set by formal risk assessment",
        "STAT", "Risk-assessed", 0,
        "ACTION REQUIRED: complete PAT risk assessment. "
        "Industrial/warehouse environments typically warrant annual or more frequent testing "
        "due to higher risk of cable damage and equipment misuse.",
    ),
]

N_COMP_COLS = len(COMP_HEADERS)
MAX_DATA_ROWS = 500   # reserve for future data entry


def build_compliance_schedule():
    wb = Workbook()
    ws = wb.active
    ws.title = "Compliance Register"

    # ── Column widths ─────────────────────────────────────────────────────────
    _set_col_widths(ws, COMP_COL_WIDTHS)

    # ── Header row ────────────────────────────────────────────────────────────
    for c, h in enumerate(COMP_HEADERS, start=1):
        ws.cell(1, c).value = h
    _style_header(ws, 1, N_COMP_COLS)
    ws.row_dimensions[1].height = 36

    # ── Data rows ─────────────────────────────────────────────────────────────
    for r, row in enumerate(COMP_DATA, start=2):
        (rec_id, site, asset_id, asset_sys, location,
         regulation, requirement, category, frequency,
         interval_days, notes) = row

        ws.cell(r, 1).value  = rec_id
        ws.cell(r, 2).value  = site
        ws.cell(r, 3).value  = asset_id
        ws.cell(r, 4).value  = asset_sys
        ws.cell(r, 5).value  = location
        ws.cell(r, 6).value  = regulation
        ws.cell(r, 7).value  = requirement
        ws.cell(r, 8).value  = category
        ws.cell(r, 9).value  = frequency
        ws.cell(r, 10).value = interval_days if interval_days > 0 else None
        # K (11): Last Completed  - blank
        # L (12): Certificate Ref - blank
        # M (13): Next Due Date
        ws.cell(r, 13).value = (
            f'=IF(OR(K{r}="",J{r}=0),"",K{r}+J{r})'
        )
        ws.cell(r, 13).number_format = "DD/MM/YYYY"
        # N (14): Days Until Due
        ws.cell(r, 14).value = f'=IF(M{r}="","",M{r}-TODAY())'
        ws.cell(r, 14).number_format = "0"
        # O (15): Status
        ws.cell(r, 15).value = (
            f'=IF(M{r}="","Not Scheduled",'
            f'IF(N{r}="","Not Scheduled",'
            f'IF(N{r}<0,"OVERDUE",'
            f'IF(N{r}<=30,"Due Within 30 Days",'
            f'IF(N{r}<=90,"Due Within 90 Days","Current")))))'
        )
        # P (16): Responsible Contractor - blank
        ws.cell(r, 17).value = notes

    # ── Style all rows ────────────────────────────────────────────────────────
    _style_data_rows(ws, 2, 2 + len(COMP_DATA) - 1, N_COMP_COLS)

    # ── Style formula cells (date-format K and M columns) ────────────────────
    for r in range(2, 2 + len(COMP_DATA)):
        ws.cell(r, 11).number_format = "DD/MM/YYYY"   # K Last Completed
        ws.cell(r, 13).number_format = "DD/MM/YYYY"   # M Next Due Date

    # ── Blank rows for future entries ─────────────────────────────────────────
    blank_start = 2 + len(COMP_DATA)
    blank_end   = blank_start + MAX_DATA_ROWS - len(COMP_DATA)
    for r in range(blank_start, blank_end + 1):
        ws.cell(r, 13).value = f'=IF(OR(K{r}="",J{r}=0),"",K{r}+J{r})'
        ws.cell(r, 13).number_format = "DD/MM/YYYY"
        ws.cell(r, 14).value = f'=IF(M{r}="","",M{r}-TODAY())'
        ws.cell(r, 14).number_format = "0"
        ws.cell(r, 15).value = (
            f'=IF(M{r}="","Not Scheduled",'
            f'IF(N{r}="","Not Scheduled",'
            f'IF(N{r}<0,"OVERDUE",'
            f'IF(N{r}<=30,"Due Within 30 Days",'
            f'IF(N{r}<=90,"Due Within 90 Days","Current")))))'
        )
        for c in range(1, N_COMP_COLS + 1):
            cell = ws.cell(r, c)
            cell.font  = BODY_FONT
            cell.border = _thin_border()
            cell.alignment = WRAP_TOP
        ws.cell(r, 11).number_format = "DD/MM/YYYY"
        ws.cell(r, 13).number_format = "DD/MM/YYYY"

    # ── Data validation dropdowns ─────────────────────────────────────────────
    _add_dv_list(ws, '"CITY,MFG"',       f"B2:B{blank_end}", "Select site", "Site")
    _add_dv_list(ws, '"STAT,PPM"',       f"H2:H{blank_end}", "STAT = Statutory/Regulatory  |  PPM = Planned Preventive", "Category")

    # ── Conditional formatting - Status column (O) ────────────────────────────
    status_range = f"O2:O{blank_end}"

    def _cf(formula, fill, bold=False):
        dxf = DifferentialStyle(
            fill=fill,
            font=Font(name="Calibri", bold=bold, size=10,
                      color="FFFFFF" if bold else "000000"),
        )
        return Rule(type="expression", formula=[formula], dxf=dxf, stopIfTrue=True)

    ws.conditional_formatting.add(status_range, _cf(f'$O2="OVERDUE"',            F_OVERDUE,  bold=True))
    ws.conditional_formatting.add(status_range, _cf(f'$O2="Due Within 30 Days"', F_DUE30,    bold=True))
    ws.conditional_formatting.add(status_range, _cf(f'$O2="Due Within 90 Days"', F_DUE90))
    ws.conditional_formatting.add(status_range, _cf(f'$O2="Current"',            F_CURRENT))
    ws.conditional_formatting.add(status_range, _cf(f'$O2="Not Scheduled"',      F_NOTSCHED))

    # Highlight HIGH RISK notes rows (Q column contains "** HIGH RISK")
    notes_range = f"A2:Q{blank_end}"
    dxf_highrisk = DifferentialStyle(fill=_fill("FFF2CC"))   # pale yellow
    ws.conditional_formatting.add(
        notes_range,
        Rule(type="expression", formula=[f'ISNUMBER(SEARCH("HIGH RISK",$Q2))'], dxf=dxf_highrisk)
    )

    # ── Freeze panes & auto-filter ────────────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:Q{blank_end}"

    # ── Row heights for data rows ──────────────────────────────────────────────
    for r in range(2, 2 + len(COMP_DATA)):
        ws.row_dimensions[r].height = 60

    # ── Instructions sheet ────────────────────────────────────────────────────
    _add_instructions_sheet(wb, "Instructions", [
        "HOW TO USE THE COMPLIANCE REGISTER",
        "",
        "SETUP (one-time)",
        "  1. For each row, enter the date the requirement was LAST completed in column K (Last Completed).",
        "  2. Column M (Next Due Date) will calculate automatically for fixed-interval items.",
        "     For 'Risk-assessed' items (no interval), enter the Next Due Date manually in column M.",
        "  3. Enter the responsible contractor in column P.",
        "",
        "ONGOING USE",
        "  4. When a job is completed, update column K (Last Completed) with the new completion date.",
        "     Column M will recalculate automatically.",
        "  5. Enter the certificate or report reference number in column L.",
        "  6. The Status column (O) updates automatically:",
        "       OVERDUE          = past due date           (red)",
        "       Due Within 30 Days = due in <= 30 days     (dark orange)",
        "       Due Within 90 Days = due in <= 90 days     (amber)",
        "       Current          = more than 90 days away  (green)",
        "       Not Scheduled    = no due date set yet     (grey)",
        "",
        "RAISING JOBS",
        "  7. When a job needs to be raised (e.g. annual boiler service is due), create a",
        "     corresponding entry in the Job Log workbook and record the Job ID in the Notes column here.",
        "",
        "ADDING NEW ASSETS",
        "  8. Add new rows at the bottom of the data. Use the next available COMP-XXXX Record ID.",
        "  9. Formulas in columns M, N and O will already be in place for blank rows.",
        "",
        "ACTIONS REQUIRED",
        "  Items marked ** ACTION REQUIRED ** in the Notes column need attention before the system",
        "  can be considered complete (e.g. confirm F-Gas CO2e charges, obtain Written Schemes).",
        "",
        "LEGIONELLA HIGH RISK ITEMS",
        "  Rows highlighted in yellow contain ** HIGH RISK ** notes.",
        "  The MFG site shower facilities require particularly close attention - see notes for each row.",
    ])

    path = os.path.join(OUTPUT_DIR, "Compliance_Schedule.xlsx")
    wb.save(path)
    print(f"Saved: {path}")


# ══════════════════════════════════════════════════════════════════════════════
#  JOB LOG
# ══════════════════════════════════════════════════════════════════════════════

JOB_HEADERS = [
    "Job ID",                  # A
    "Raised Date",             # B
    "Site",                    # C
    "Asset ID",                # D
    "Asset / Location Desc.",  # E
    "Category",                # F
    "Regulation / Standard",   # G
    "Compliance Record ID",    # H  links to COMP-XXXX
    "Description of Work",     # I
    "Priority",                # J
    "Assigned Contractor",     # K
    "Status",                  # L
    "Target Completion",       # M
    "Actual Completion",       # N
    "Quoted Cost (£)",         # O
    "Actual Cost (£)",         # P
    "Invoice Reference",       # Q
    "Certificate Received?",   # R
    "Compliance Updated?",     # S
    "Closed By",               # T
    "Notes",                   # U
]

JOB_COL_WIDTHS = {
    "A": 11, "B": 13, "C":  8, "D": 16, "E": 28,
    "F": 10, "G": 28, "H": 15, "I": 38, "J": 11,
    "K": 24, "L": 18, "M": 14, "N": 14, "O": 14,
    "P": 14, "Q": 18, "R": 16, "S": 16, "T": 16,
    "U": 38,
}

N_JOB_COLS = len(JOB_HEADERS)
JOB_MAX_ROWS = 2000


def build_job_log():
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Log"

    # ── Column widths ─────────────────────────────────────────────────────────
    _set_col_widths(ws, JOB_COL_WIDTHS)

    # ── Header row ────────────────────────────────────────────────────────────
    for c, h in enumerate(JOB_HEADERS, start=1):
        ws.cell(1, c).value = h
    _style_header(ws, 1, N_JOB_COLS)
    ws.row_dimensions[1].height = 36

    # ── Style blank data rows ─────────────────────────────────────────────────
    for r in range(2, JOB_MAX_ROWS + 2):
        fill = BLUE_LIGHT if r % 2 == 0 else WHITE_FILL
        for c in range(1, N_JOB_COLS + 1):
            cell = ws.cell(r, c)
            cell.fill  = fill
            cell.font  = BODY_FONT
            cell.alignment = WRAP_TOP
            cell.border = _thin_border()
        ws.cell(r,  2).number_format = "DD/MM/YYYY"   # Raised Date
        ws.cell(r, 13).number_format = "DD/MM/YYYY"   # Target Completion
        ws.cell(r, 14).number_format = "DD/MM/YYYY"   # Actual Completion
        ws.cell(r, 15).number_format = '£#,##0.00'    # Quoted Cost
        ws.cell(r, 16).number_format = '£#,##0.00'    # Actual Cost

    # ── Data validation dropdowns ─────────────────────────────────────────────
    last = JOB_MAX_ROWS + 1
    _add_dv_list(ws, '"CITY,MFG"',
                 f"C2:C{last}",
                 "Select site code", "Site")
    _add_dv_list(ws, '"STAT,PPM,REACT,EMERG"',
                 f"F2:F{last}",
                 "STAT=Statutory | PPM=Planned Preventive | REACT=Reactive | EMERG=Emergency",
                 "Category")
    _add_dv_list(ws, '"Critical,High,Medium,Low"',
                 f"J2:J{last}",
                 "Critical = immediate safety risk | High = urgent | Medium = routine | Low = desirable",
                 "Priority")
    _add_dv_list(ws, '"Open,In Progress,Pending Invoice,Closed,Cancelled"',
                 f"L2:L{last}",
                 "Select current job status", "Status")
    _add_dv_list(ws, '"Yes,No,N/A"',
                 f"R2:R{last}",
                 "Has the certificate or report been received and filed?",
                 "Certificate Received?")
    _add_dv_list(ws, '"Yes,No,N/A"',
                 f"S2:S{last}",
                 "Has the Compliance Register been updated with the new completion date?",
                 "Compliance Updated?")

    # ── Conditional formatting - Status column (L) ────────────────────────────
    status_range = f"L2:L{JOB_MAX_ROWS + 1}"

    def _cf_job(formula, fill, bold=False, text_color="000000"):
        dxf = DifferentialStyle(
            fill=fill,
            font=Font(name="Calibri", bold=bold, size=10, color=text_color),
        )
        return Rule(type="expression", formula=[formula], dxf=dxf, stopIfTrue=True)

    ws.conditional_formatting.add(status_range, _cf_job('$L2="Open"',            F_OPEN))
    ws.conditional_formatting.add(status_range, _cf_job('$L2="In Progress"',     F_INPROG))
    ws.conditional_formatting.add(status_range, _cf_job('$L2="Pending Invoice"', F_PEND_INV))
    ws.conditional_formatting.add(status_range, _cf_job('$L2="Closed"',          F_CLOSED))
    ws.conditional_formatting.add(status_range, _cf_job('$L2="Cancelled"',       F_CANCELLED))

    # Highlight EMERG rows across full row
    full_range = f"A2:U{JOB_MAX_ROWS + 1}"
    dxf_emerg = DifferentialStyle(fill=_fill("FFE0E0"))   # pale red
    ws.conditional_formatting.add(
        full_range,
        Rule(type="expression", formula=['$F2="EMERG"'], dxf=dxf_emerg)
    )

    # Highlight overdue open/in-progress jobs (target date past, not closed)
    dxf_overdue = DifferentialStyle(
        fill=_fill("FF0000"),
        font=Font(name="Calibri", bold=True, size=10, color="FFFFFF"),
    )
    ws.conditional_formatting.add(
        f"M2:M{JOB_MAX_ROWS + 1}",
        Rule(
            type="expression",
            formula=['AND($M2<TODAY(),OR($L2="Open",$L2="In Progress"))'],
            dxf=dxf_overdue,
        )
    )

    # ── Freeze panes & auto-filter ────────────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:U{JOB_MAX_ROWS + 1}"

    # ── Instructions sheet ────────────────────────────────────────────────────
    _add_instructions_sheet(wb, "Instructions", [
        "HOW TO USE THE JOB LOG",
        "",
        "RAISING A NEW JOB",
        "  1. Enter the next Job ID in column A (format: JOB-NNNN, e.g. JOB-0001, JOB-0042).",
        "  2. Enter today's date in column B (Raised Date).",
        "  3. Select the site (CITY / MFG) in column C.",
        "  4. Enter the Asset ID from the Compliance Register or Asset Register in column D (if known).",
        "  5. Enter a brief description of the asset or location in column E.",
        "  6. Select the job category in column F:",
        "       STAT  = Statutory/Regulatory requirement",
        "       PPM   = Planned Preventive Maintenance",
        "       REACT = Reactive (reported defect, planned response)",
        "       EMERG = Emergency (immediate safety-critical response) - row highlighted in pale red",
        "  7. If STAT, enter the regulation name in column G and the Compliance Record ID (COMP-XXXX)",
        "     from the Compliance Register in column H.",
        "  8. Enter the description of work required in column I.",
        "  9. Select priority in column J and enter the target completion date in column M.",
        "     Target date highlighted red if overdue and job is still Open or In Progress.",
        "",
        "PROGRESSING A JOB",
        " 10. Update Status (column L) as the job progresses:",
        "       Open > In Progress > Pending Invoice > Closed",
        " 11. Enter the contractor name in column K when assigned.",
        " 12. Enter quoted cost in column O when received.",
        "",
        "CLOSING A JOB",
        " 13. Enter actual completion date (column N) and actual cost (column P).",
        " 14. Enter invoice reference in column Q.",
        " 15. Mark Certificate Received? (column R) - Yes / No / N/A.",
        " 16. Mark Compliance Updated? (column S) - go to Compliance Register and update",
        "     the Last Completed date and Certificate Ref for the corresponding COMP record.",
        " 17. Set Status to Closed and enter your name in Closed By (column T).",
        "",
        "JOB ID NUMBERING",
        "  Use the format JOB-NNNN with leading zeros (JOB-0001, JOB-0002 ... JOB-0999, JOB-1000).",
        "  Never reuse a Job ID. Cancelled jobs retain their ID with Status = Cancelled.",
        "",
        "COST TRACKING",
        "  Columns O (Quoted) and P (Actual) are formatted as GBP currency.",
        "  Enter numeric values only - do not type the £ sign.",
        "  Use Excel pivot tables or the =SUMIFS formula to report spend by site, category or contractor.",
    ])

    path = os.path.join(OUTPUT_DIR, "Job_Log.xlsx")
    wb.save(path)
    print(f"Saved: {path}")


# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    build_compliance_schedule()
    build_job_log()
    print("Done.")
